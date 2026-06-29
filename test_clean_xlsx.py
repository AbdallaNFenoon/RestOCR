import os
import openpyxl
import re

def clean_formulas_in_wb(wb):
    print("Scanning and cleaning UDF formulas...")
    cleaned_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    if '__xludf.DUMMYFUNCTION' in formula:
                        pattern = r'__xludf\.DUMMYFUNCTION\("((?:[^"]|"")*)"\)'
                        match = re.search(pattern, formula)
                        if match:
                            inner = match.group(1).replace('""', '"')
                            cleaned = formula.replace(match.group(0), inner)
                            cell.value = cleaned
                            cleaned_count += 1
    print(f"Cleaned {cleaned_count} formula cells")

def main():
    template_path = os.path.join("backend", "template.xlsx")
    wb = openpyxl.load_workbook(template_path)
    clean_formulas_in_wb(wb)
    wb.save("cleaned_template.xlsx")
    
    # Let's inspect some formulas in the cleaned workbook
    wb2 = openpyxl.load_workbook("cleaned_template.xlsx")
    ws = wb2['item']
    print("\nRow 2 image formula after cleaning:")
    print(ws.cell(row=2, column=8).value)

if __name__ == "__main__":
    main()
