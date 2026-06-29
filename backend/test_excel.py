import os
import openpyxl
from backend.excel_service import generate_excel

def run_test():
    print("=== Testing Excel Generation ===")
    
    mock_menu = {
        "categories": [
            {
                "name_en": "Fresh Karahi",
                "name_ar": "كراهي طازجة",
                "average_prep_time_minutes": 25,
                "is_active": 1,
                "items": [
                    {
                        "name_en": "Chicken Karahi",
                        "name_ar": "كراهي دجاج",
                        "price": 0.0,
                        "options": [
                            {"name_en": "Half", "name_ar": "نصف", "price": 30.0},
                            {"name_en": "Full", "name_ar": "كامل", "price": 54.0}
                        ],
                        "tags": ["non_veg"],
                        "item_type": "non_veg"
                    },
                    {
                        "name_en": "Chicken Boneless Handi",
                        "name_ar": "هاندي دجاج بدون عظم",
                        "price": 24.0,
                        "options": [],
                        "tags": ["non_veg"],
                        "item_type": "non_veg"
                    }
                ]
            },
            {
                "name_en": "Daal",
                "name_ar": "دال",
                "average_prep_time_minutes": 15,
                "is_active": 1,
                "items": [
                    {
                        "name_en": "Shahi Daal",
                        "name_ar": "دال شاهي",
                        "price": 15.0,
                        "options": [],
                        "tags": ["veg"],
                        "item_type": "veg"
                    }
                ]
            }
        ]
    }
    
    template_path = os.path.join("backend", "template.xlsx")
    output_path = "test_output.xlsx"
    
    if not os.path.exists(template_path):
        print(f"Error: Template not found at {template_path}")
        return False
        
    try:
        generate_excel(mock_menu, template_path, output_path)
        print(f"Excel generated successfully at {output_path}")
        
        # Load and verify contents
        wb = openpyxl.load_workbook(output_path)
        
        print("\nVerifying sheets:")
        for sheet_name in ['category', 'item', 'modifier_group', 'modifier_item']:
            if sheet_name not in wb.sheetnames:
                print(f"  [FAIL] Sheet '{sheet_name}' is missing!")
                return False
            
            ws = wb[sheet_name]
            print(f"  Sheet '{sheet_name}': {ws.max_row} rows (including headers)")
            
            # Print row data safely without crashing on encoding
            for r in range(2, ws.max_row + 1):
                row_vals = [cell.value for cell in ws[r]]
                # Encode values to ascii with backslashreplace for safe printing
                safe_row_vals = [str(val).encode('ascii', errors='backslashreplace').decode('ascii') for val in row_vals]
                print(f"    Row {r}: {safe_row_vals}")
                
        wb.close()
        print("\n[SUCCESS] Excel generation test passed!")
        return True
        
    except Exception as e:
        print(f"\n[FAIL] Test failed with error: {e}")
        return False

if __name__ == "__main__":
    run_test()
