import openpyxl
import os
import re
import io
import zipfile
import logging
from typing import Dict, List, Any
from copy import copy

logger = logging.getLogger(__name__)

# Column mappings as they appear in the template sheets (in exact order)
CATEGORY_COLS = [
    'category_code', 'name_en', 'name_ar', 'average_prep_time_minutes', 
    'external_schedule_code', 'internal_category_code', 'is_active'
]

ITEM_COLS = [
    'item_code', 'category_code', 'name_en', 'name_ar', 'description_en', 
    'description_ar', 'price', 'image', 'modifier_groups', 'tags', 
    'calories', 'item_type', 'external_schedule_code', 'fake_original_price', 
    'max_qty', 'in_stock', 'is_active'
]

MODIFIER_GROUP_COLS = [
    'modifier_group_code', 'name_en', 'name_ar', 'description_en', 
    'description_ar', 'max_quantity_per_item', 'max_total_selections', 
    'min_total_selections', 'modifier_items', 'is_active'
]

MODIFIER_ITEM_COLS = [
    'modifier_item_code', 'name_en', 'name_ar', 'description_en', 
    'description_ar', 'price', 'image', 'item_code', 'modifier_groups', 
    'calories', 'modifier_type', 'in_stock', 'is_active'
]

MODIFIER_GROUP_OPTION_COLS = [
    'modifier_group_code', 'modifier_item_code', 'overriden_price'
]

WRITE_COLS = {
    'category': [
        'category_code', 'name_en', 'name_ar', 'external_schedule_code', 'internal_category_code'
    ],
    'item': [
        'item_code', 'category_code', 'name_en', 'name_ar', 'description_en', 
        'description_ar', 'price', 'modifier_groups', 'calories', 'item_type', 
        'external_schedule_code', 'fake_original_price', 'max_qty', 'in_stock'
    ],
    'modifier_group': [
        'modifier_group_code', 'name_en', 'name_ar', 'description_ar', 
        'max_quantity_per_item', 'max_total_selections', 'min_total_selections', 'modifier_items'
    ],
    'modifier_item': [
        'modifier_item_code', 'name_en', 'name_ar', 'description_ar', 
        'price', 'image', 'modifier_groups', 'calories', 'modifier_type', 'in_stock'
    ],
    'modifier_group_option': [
        'modifier_group_code', 'modifier_item_code', 'overriden_price'
    ]
}

def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        if isinstance(val, str):
            val = val.replace('$', '').replace('SAR', '').replace('SR', '').strip()
            if not val:
                return default
        return float(val)
    except (ValueError, TypeError):
        return default

def safe_int(val, default=20):
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default

def clean_all_formulas(wb):
    logger.info("Scanning and removing all formulas in workbook to prevent Excel corruption...")
    cleaned_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    cell.value = None
                    cleaned_count += 1
    logger.info(f"Removed {cleaned_count} formulas to ensure Excel compatibility.")

def generate_excel_in_memory(menu_data: Dict[str, Any], template_path: str) -> bytes:
    """
    Loads template_path, clears existing rows, populates them with menu_data,
    and returns the workbook as bytes in memory.
    """
    logger.info(f"Generating Excel in memory from template {template_path}")
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template Excel file not found at {template_path}")
        
    wb = openpyxl.load_workbook(template_path)
    clean_all_formulas(wb)
    
    # Process categories
    categories = []
    items = []
    modifier_groups = []
    modifier_items = []
    modifier_group_options = []
    
    cat_id_counter = 1
    item_id_counter = 1
    mod_grp_counter = 1
    mod_item_counter = 1
    
    for cat_idx, cat in enumerate(menu_data.get('categories', [])):
        cat_code = f"C{cat_id_counter}"
        cat_id_counter += 1
        
        categories.append({
            'category_code': cat_code,
            'name_en': cat.get('name_en', '').strip(),
            'name_ar': cat.get('name_ar', '').strip(),
            'average_prep_time_minutes': safe_int(cat.get('average_prep_time_minutes'), 20),
            'external_schedule_code': cat.get('external_schedule_code', None),
            'internal_category_code': cat.get('internal_category_code', None),
            'is_active': cat.get('is_active', 1)
        })
        
        for item in cat.get('items', []):
            item_code = f"I{item_id_counter}"
            item_id_counter += 1
            
            # Check if this item has multiple options/sizes
            options = item.get('options', [])
            mod_grp_code = None
            
            if len(options) >= 2:
                # Multiple options (e.g. Half/Full): base price is 0.0, attach a modifier group
                item_price = 0.0
                mod_grp_code = f"M{mod_grp_counter}"
                mod_grp_counter += 1
                
                # Create modifier items and record their codes
                opt_codes = []
                for opt in options:
                    opt_code = f"O{mod_item_counter}"
                    mod_item_counter += 1
                    opt_codes.append(opt_code)
                    
                    modifier_items.append({
                        'modifier_item_code': opt_code,
                        'name_en': str(opt.get('name_en', '')).strip(),
                        'name_ar': str(opt.get('name_ar', '')).strip(),
                        'description_en': None,
                        'description_ar': None,
                        'price': safe_float(opt.get('price')),
                        'image': None,
                        'item_code': None,
                        'modifier_groups': None,
                        'calories': None,
                        'modifier_type': None,
                        'in_stock': 1,
                        'is_active': 1
                    })
                    
                    # Link modifier options to group
                    modifier_group_options.append({
                        'modifier_group_code': mod_grp_code,
                        'modifier_item_code': opt_code,
                        'overriden_price': None
                    })
                
                # Create the modifier group mapping to these items
                modifier_groups.append({
                    'modifier_group_code': mod_grp_code,
                    'name_en': 'Choose',
                    'name_ar': 'اختر',
                    'description_en': None,
                    'description_ar': None,
                    'max_quantity_per_item': 1,
                    'max_total_selections': 1,
                    'min_total_selections': 1,
                    'modifier_items': ','.join(opt_codes),
                    'is_active': 1
                })
            elif len(options) == 1:
                # Single option: collapse to a direct price — no modifier group needed.
                # This handles items like "Chicken Boneless Handi (Full only)" → just price=24.
                item_price = safe_float(options[0].get('price'))
            else:
                # No options: use direct price from the item's price field
                item_price = safe_float(item.get('price'))
            
            # Map tags list to comma-separated string or use default
            tags = item.get('tags', [])
            tags_str = ','.join(tags) if isinstance(tags, list) else str(tags)
            
            items.append({
                'item_code': item_code,
                'category_code': cat_code,
                'name_en': item.get('name_en', '').strip(),
                'name_ar': item.get('name_ar', '').strip(),
                'description_en': item.get('description_en', None),
                'description_ar': item.get('description_ar', None),
                'price': item_price,
                'image': None,
                'modifier_groups': mod_grp_code,
                'tags': tags_str if tags_str else None,
                'calories': item.get('calories', None),
                'item_type': item.get('item_type', 'non_veg'),
                'external_schedule_code': item.get('external_schedule_code', None),
                'fake_original_price': item.get('fake_original_price', None),
                'max_qty': item.get('max_qty', None),
                'in_stock': 1,
                'is_active': 1
            })

    def write_to_sheet(sheet_name: str, all_cols: List[str], data: List[Dict[str, Any]]):
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet {sheet_name} not found in template. Skipping.")
            return
        ws = wb[sheet_name]
        
        write_cols = WRITE_COLS.get(sheet_name, all_cols)
        current_max_row = ws.max_row
                
        # Write new data starting exactly at row 2 using explicit cell coordinates
        for r_idx, data_row in enumerate(data):
            row_num = 2 + r_idx
            
            # If we are writing past current_max_row, copy styles and formulas from row 2
            if row_num > current_max_row:
                for col_idx, col in enumerate(all_cols):
                    cell_src = ws.cell(row=2, column=col_idx + 1)
                    new_cell = ws.cell(row=row_num, column=col_idx + 1)
                    
                    # Copy styles
                    if cell_src.has_style:
                        new_cell.font = copy(cell_src.font)
                        new_cell.border = copy(cell_src.border)
                        new_cell.fill = copy(cell_src.fill)
                        new_cell.number_format = copy(cell_src.number_format)
                        new_cell.protection = copy(cell_src.protection)
                        new_cell.alignment = copy(cell_src.alignment)

                    # Copy formulas for columns we don't write to
                    if col not in write_cols:
                        if isinstance(cell_src.value, str) and cell_src.value.startswith('='):
                            import re
                            formula = cell_src.value
                            def adjust_formula_cell(match):
                                if match.group(1):
                                    return match.group(1)
                                else:
                                    col_abs = match.group(2) or ''
                                    col_letter = match.group(3)
                                    row_abs = match.group(4) or ''
                                    return f"{col_abs}{col_letter}{row_abs}{row_num}"
                            
                            adjusted_formula = re.sub(
                                r'("[^"\\]*(?:\\.[^"\\]*)*")|(\$?)([BC])(\$?)(2)\b',
                                adjust_formula_cell,
                                formula
                            )
                            ws.cell(row=row_num, column=col_idx + 1, value=adjusted_formula)
            
            # Write data values
            for col in write_cols:
                col_idx = all_cols.index(col)
                val = data_row.get(col, None)
                ws.cell(row=row_num, column=col_idx + 1, value=val)
                
        # Clear values in target columns for remaining rows to clean up leftovers
        for r_idx in range(len(data), ws.max_row - 1):
            row_num = 2 + r_idx
            for col in write_cols:
                col_idx = all_cols.index(col)
                ws.cell(row=row_num, column=col_idx + 1, value=None)

    # Populate sheets
    write_to_sheet('category', CATEGORY_COLS, categories)
    write_to_sheet('item', ITEM_COLS, items)
    write_to_sheet('modifier_group', MODIFIER_GROUP_COLS, modifier_groups)
    write_to_sheet('modifier_item', MODIFIER_ITEM_COLS, modifier_items)
    write_to_sheet('modifier_group_option', MODIFIER_GROUP_OPTION_COLS, modifier_group_options)
    
    # Schedule sheet is left untouched as it contains pre-configured schedules.
    
    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    wb.close()
    
    # Post-process: remove Google-specific XML content that causes Excel warnings.
    # openpyxl preserves these artifacts from the template even after patching.
    file_bytes = _clean_google_artifacts(virtual_file.getvalue())
    logger.info("Excel file generated successfully in memory!")
    return file_bytes

def _clean_google_artifacts(file_bytes: bytes) -> bytes:
    """
    Post-processes the xlsx bytes to strip Google Sheets-specific content
    that makes Excel show a 'problem with content' recovery warning:
      - Google-specific fonts (Google Sans Mono, Docs-Calibri) in styles.xml
      - Google-specific named ranges (Z_<GUID>_...) in workbook.xml
    """
    buffer = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as z_in:
        with zipfile.ZipFile(buffer, 'w', compression=zipfile.ZIP_DEFLATED) as z_out:
            for name in z_in.namelist():
                data = z_in.read(name)
                
                if name == 'xl/styles.xml':
                    text = data.decode('utf-8')
                    # Replace Google-specific font names with standard equivalents.
                    # openpyxl stores them as &quot;docs-Google Sans Mono&quot; (with docs- prefix).
                    # Use regex to catch all variants (with or without docs- prefix, any casing).
                    text = re.sub(
                        r'(&quot;|")docs-Google Sans Mono(&quot;|")',
                        'Consolas',
                        text,
                        flags=re.IGNORECASE
                    )
                    text = re.sub(
                        r'(&quot;|")Google Sans Mono(&quot;|")',
                        'Consolas',
                        text,
                        flags=re.IGNORECASE
                    )
                    text = text.replace('Docs-Calibri', 'Calibri')
                    data = text.encode('utf-8')
                    
                elif name == 'xl/workbook.xml':
                    text = data.decode('utf-8')
                    # Remove Google Sheets internal named ranges (Z_<GUID>_ pattern)
                    text = re.sub(
                        r'<definedName[^>]*name="Z_[^"]*"[^/]*/?>',
                        '',
                        text
                    )
                    text = re.sub(
                        r'<definedName[^>]*name="Z_[^"]*"[^>]*>.*?</definedName>',
                        '',
                        text,
                        flags=re.DOTALL
                    )
                    # Remove empty definedNames block if all entries were removed
                    text = re.sub(r'<definedNames>\s*</definedNames>', '', text)
                    data = text.encode('utf-8')
                    
                z_out.writestr(name, data)
    return buffer.getvalue()


def generate_excel(menu_data: Dict[str, Any], template_path: str, output_path: str):
    """Wrapper to write in-memory generated workbook to disk (used for testing)."""
    file_bytes = generate_excel_in_memory(menu_data, template_path)
    with open(output_path, "wb") as f:
        f.write(file_bytes)
