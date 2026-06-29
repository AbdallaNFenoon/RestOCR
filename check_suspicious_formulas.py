import openpyxl

wb = openpyxl.load_workbook('test_output_final.xlsx')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r in range(2, 5):
        for c in range(1, 20):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                val = cell.value
                if '()' in val or ',)' in val or '(,' in val:
                    print(f"SUSPICIOUS FORMULA in {sheet_name} {cell.coordinate}: {val}")
